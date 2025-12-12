"""
File Processors for the Anonymization Tool

This module uses a Template Method Pattern. The `FileProcessor` base class defines
the main workflow for processing files, and subclasses implement the specific
details for extracting text from different file formats (e.g., PDF, DOCX, JSON).
"""
import copy
import gc
import io
import os
import ijson
import re
import shutil
import logging
import unicodedata 
from abc import ABC, abstractmethod
from collections import defaultdict, deque
from typing import Iterable, List, Generator, Optional, Dict, Tuple, Union
from pathlib import Path # Added Path import

import openpyxl  # type: ignore
import orjson  # type: ignore
import pandas as pd  # type: ignore
import pytesseract  # type: ignore
from docx import Document  # type: ignore
from lxml import etree  # type: ignore
from PIL import Image  # type: ignore
from tqdm.auto import tqdm  # type: ignore
import pymupdf as fitz  # type: ignore

from .config import Global, ProcessingLimits, DefaultSizes
from .engine import AnonymizationOrchestrator

def get_output_path(original_path: str, new_ext: str, prefix: str = "anon_", output_dir: str = "output") -> str:
    """Constructs a secure output file path, preventing path traversal."""
    
    # 1. Resolve and validate the output directory path
    project_root = Path(os.getcwd()).resolve()
    output_path_obj = Path(output_dir).resolve()
    
    # Ensure the resolved output directory is inside the project root
    if not output_path_obj.is_relative_to(project_root):
        logging.error(f"Path traversal attempt detected: Output directory '{output_dir}' is outside the project boundary '{project_root}'.")
        raise ValueError(f"Path traversal attempt detected: Output directory '{output_dir}' is outside the project boundary.")

    output_path_obj.mkdir(parents=True, exist_ok=True)
    
    # 2. Sanitize filename from original_path to prevent it from being used for traversal
    original_base_name = Path(original_path).name # Get only the filename from the path
    
    # Normalize Unicode (NFD) and remove dangerous characters for robust sanitization
    # Allow only alphanumeric, hyphen, underscore, and a single dot for extension
    sanitized_base_name = unicodedata.normalize('NFKD', original_base_name).encode('ascii', 'ignore').decode('utf-8')
    
    # Separate base name and extension to sanitize them individually
    name_part, ext_part = os.path.splitext(sanitized_base_name)
    
    # Sanitize the name part: remove any character that is not alphanumeric or hyphen/underscore
    name_part = re.sub(r'[^\w\-]', '', name_part)
    
    # Sanitize the extension part (optional, as new_ext will overwrite it anyway)
    # For extra safety, ensure ext_part does not contain path separators or dangerous chars
    ext_part = re.sub(r'[^\w\.]', '', ext_part) # Keep only alphanumeric and dot for extension

    # Reconstruct a safe base name; new_ext will be the final extension
    safe_base_name = f"{name_part}{ext_part}" 

    if not safe_base_name or safe_base_name in (".", ".."):
        logging.error(f"Invalid or sanitized-away filename from original path: '{original_path}' resulted in an empty or invalid name.")
        raise ValueError(f"Invalid filename derived from original path: '{original_path}'")

    # Construct the final filename: prefix + sanitized_name + new_ext
    final_filename = f"{prefix}{name_part}{new_ext}"
    
    # 3. Construct the full candidate path and perform the final check
    candidate_file_path = output_path_obj / final_filename
    resolved_candidate_file_path = candidate_file_path.resolve()
    
    # Final check to ensure the candidate path is within the resolved output directory.
    if not resolved_candidate_file_path.is_relative_to(output_path_obj):
        logging.error(f"Path traversal attempt detected for final file path: '{resolved_candidate_file_path}' is outside output directory '{output_path_obj}'.")
        raise ValueError(f"Path traversal attempt detected for final file path: '{resolved_candidate_file_path}'.")
        
    return str(resolved_candidate_file_path)


def extract_text_from_image(image_bytes: bytes) -> str:
    """Extracts text from image bytes using OCR."""
    try:
        with Image.open(io.BytesIO(image_bytes)) as img:
            return pytesseract.image_to_string(img)
    except Exception as e:
        logging.exception(f"Error during OCR extraction: {e}")
        return ""


class FileProcessor(ABC):
    DEFAULT_BATCH_SIZE = 200

    def __init__(self, file_path: str, orchestrator: AnonymizationOrchestrator, 
                 ner_data_generation: bool = False, 
                 anonymization_config: Optional[Dict] = None, 
                 min_word_length: int = 3, 
                 skip_numeric: bool = False, 
                 output_dir: str = "output", 
                 overwrite: bool = False, 
                 disable_gc: bool = False,
                 json_stream_threshold_mb: int = 100,
                 preserve_row_context: bool = False,
                 batch_size: int = 200,
                 csv_chunk_size: int = 1000,
                 json_chunk_size: int = 1000,
                 ner_chunk_size: int = 1500,
                 force_large_xml: bool = False): # Added force_large_xml
        self.file_path = file_path
        self.orchestrator = orchestrator
        self.ner_data_generation = ner_data_generation
        self.anonymization_config = copy.deepcopy(anonymization_config) if anonymization_config is not None else {}
        self.min_word_length = min_word_length
        self.skip_numeric = skip_numeric
        self.output_dir = output_dir
        self.overwrite = overwrite
        self.disable_gc = disable_gc
        self.json_stream_threshold_bytes = json_stream_threshold_mb * 1024 * 1024
        self.preserve_row_context = preserve_row_context
        self.batch_size = batch_size
        self.csv_chunk_size = csv_chunk_size
        self.json_chunk_size = json_chunk_size
        self.ner_chunk_size = ner_chunk_size
        self.force_large_xml = force_large_xml # Store the new parameter
        self.ner_output_file: Optional[str] = None
        self.ner_file_handle: Optional[io.TextIOWrapper] = None # Initialize to None with proper type hint
        logging.debug(f"FileProcessor initialized for '{file_path}' with: ner_data_generation={ner_data_generation}, min_word_length={min_word_length}, skip_numeric={skip_numeric}, output_dir='{output_dir}', overwrite={overwrite}, disable_gc={disable_gc}, force_large_xml={force_large_xml}.")

    def _get_ner_output_path(self) -> str:
        return get_output_path(self.file_path, ".jsonl", prefix="ner_data_anon_", output_dir=self.output_dir)

    def _setup_optimization(self):
        if self.disable_gc:
            gc.disable()

    def _cleanup_optimization(self):
        if self.disable_gc:
            gc.collect()
            gc.enable()
        if hasattr(self, 'ner_file_handle') and self.ner_file_handle:
            self.ner_file_handle.close()

    def _batch_iterator(self, iterator: Iterable, size: int) -> Generator[List, None, None]:
        batch = []
        for item in iterator:
            batch.append(item)
            if len(batch) >= size:
                yield batch
                batch = []
        if batch:
            yield batch

    def process(self) -> str:
        self._setup_optimization()
        output_path: str = ""
        logging.info(f"Starting processing for file: {self.file_path}")
        try:
            if self.ner_data_generation:
                logging.info("Mode: NER data generation.")
                output_path = self.ner_output_file or self._get_ner_output_path()
                logging.debug(f"NER data output path: {output_path}")
                if os.path.exists(output_path) and not self.overwrite:
                    logging.warning(f"Output file '{output_path}' already exists. Use --overwrite to replace it.")
                    return output_path
                
                # Open the file handle here, within the try block
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                self.ner_file_handle = open(output_path, "w", encoding="utf-8")
                
                all_texts = self._extract_all_texts()
                self._run_ner_pipeline(all_texts)
                logging.info(f"Successfully generated NER data to: {output_path}")
            else:
                logging.info("Mode: Anonymization.")
                output_path = get_output_path(self.file_path, self._get_output_extension(), output_dir=self.output_dir)
                logging.debug(f"Anonymized output path: {output_path}")
                if os.path.exists(output_path) and not self.overwrite:
                    logging.warning(f"Output file '{output_path}' already exists. Use --overwrite to replace it.")
                    return output_path
                self._process_anonymization(output_path)
                logging.info(f"Successfully anonymized file to: {output_path}")
        finally:
            self._cleanup_optimization()
        logging.info(f"Finished processing for file: {self.file_path}")
        return output_path

    def _process_anonymization(self, output_path: str):
        with open(output_path, "w", encoding="utf-8") as outfile:
            text_iterator = self._extract_texts()
            batch_count = 0
            for text_batch in self._batch_iterator(text_iterator, self.batch_size):
                if not text_batch: continue
                batch_count += 1
                logging.debug(f"Processing batch {batch_count} for anonymization (size: {len(text_batch)}).")
                should_anonymize, forced_type = self._should_anonymize(text_batch[0], path="<batch_first_item>") # Heuristic for batch
                if should_anonymize:
                    anonymized_batch = self._process_batch_smart(text_batch, forced_entity_type=forced_type)
                    outfile.write("".join(anonymized_batch))
                    logging.debug(f"Batch {batch_count} anonymized and written.")
                else:
                    outfile.write("".join(text_batch))
                    logging.debug(f"Batch {batch_count} skipped anonymization and original content written.")


    @abstractmethod
    def _extract_texts(self) -> Iterable[str]:
        raise NotImplementedError

    def _extract_all_texts(self) -> List[str]:
        return list(self._extract_texts())

    @abstractmethod
    def _get_output_extension(self) -> str:
        raise NotImplementedError


    def _should_anonymize(self, text: str, path: str = "") -> Tuple[bool, Optional[Union[str, List[str]]]]:
        """
        Determines if a given text should be anonymized based on a rich configuration.
        This version correctly handles path matching for array elements.
        """
        logging.debug(f"Checking _should_anonymize for text: '{text[:50]}...', path: '{path}'")
        current_path = path.lstrip('.')
        # Create a "generalized" path for matching rules that don't care about list indices.
        # e.g., "asset.tags[1].value" -> "asset.tags.value"
        generalized_path = re.sub(r'\[\d+\]', '', current_path).lstrip('.')

        # --- Configuration-based logic ---
        if self.anonymization_config:
            # 1. Highest priority: explicit exclusion.
            # A rule "asset.tags" should exclude "asset.tags[0].value".
            for rule in self.anonymization_config.get('fields_to_exclude', []):
                if generalized_path == rule or generalized_path.startswith(f"{rule}."):
                    logging.debug(f"Excluded by rule '{rule}'. Not anonymizing.")
                    return False, None

            # 2. Second highest priority: forced anonymization (bypasses text filters).
            # An exact match on the specific or generalized path triggers this.
            force_config = self.anonymization_config.get('force_anonymize', {})
            if current_path in force_config:
                entity_type = force_config[current_path].get("entity_type")
                logging.debug(f"Forced anonymization by exact path match '{current_path}'. Entity type: {entity_type}. Anonymizing.")
                return True, entity_type
            if generalized_path in force_config:
                entity_type = force_config[generalized_path].get("entity_type")
                logging.debug(f"Forced anonymization by generalized path match '{generalized_path}'. Entity type: {entity_type}. Anonymizing.")
                return True, entity_type

        # --- Text-based filtering for auto-detection ---
        if not isinstance(text, str) or len(text.strip()) < self.min_word_length:
            logging.debug(f"Skipping due to text type or length ({len(text.strip())} < {self.min_word_length}). Not anonymizing.")
            return False, None
        
        stripped_text = text.strip()
        if self.skip_numeric and stripped_text.isnumeric():
            logging.debug("Skipping numeric-only string. Not anonymizing.")
            return False, None
        
        if stripped_text.lower() in Global.TECHNICAL_STOPLIST:
            logging.debug("Skipping due to technical stoplist. Not anonymizing.")
            return False, None

        # --- Mode-based logic (explicit vs. implicit) ---
        if self.anonymization_config:
            is_explicit_mode = 'force_anonymize' in self.anonymization_config or \
                               'fields_to_anonymize' in self.anonymization_config
            
            if is_explicit_mode:
                for rule in self.anonymization_config.get('fields_to_anonymize', []):
                    if generalized_path == rule or generalized_path.startswith(f"{rule}."):
                        logging.debug(f"Explicit mode: matched 'fields_to_anonymize' rule '{rule}'. Anonymizing.")
                        return True, None
                logging.debug("Explicit mode: no matching anonymization rule found. Not anonymizing.")
                return False, None

        # Default to anonymizing if it passed all checks in implicit mode.
        logging.debug("Implicit mode: no explicit rules, passed text filters. Anonymizing.")
        return True, None

    def _process_batch_smart(self, text_list: List[str], forced_entity_type: Optional[Union[str, List[str]]] = None) -> List[str]:
        if not text_list:
            logging.debug("Received empty text_list for batch processing. Returning empty.")
            return []
        
        logging.debug(f"Processing smart batch of {len(text_list)} items. Forced entity type: {forced_entity_type}")
        
        anonymized_values = self.orchestrator.anonymize_texts(
            text_list,
            forced_entity_type=forced_entity_type
        )

        if len(anonymized_values) != len(text_list):
            logging.critical(f"PII leakage detected in {self.file_path}: batch size mismatch. "
                             f"Input length: {len(text_list)}, Output length: {len(anonymized_values)}")
            raise RuntimeError("Anonymization failed to prevent data leak. Halting execution.")
        
        logging.debug(f"Smart batch processed successfully. Anonymized {len(text_list)} items.")
        return anonymized_values

    def _run_ner_pipeline(self, text_list: List[str]):
        if not text_list: return

        MAX_CHUNK_SIZE = DefaultSizes.NER_CHUNK_SIZE
        DELIMITER = " . ||| . "
        
        text_chunks: List[str] = []
        current_chunk: List[str] = []
        current_len = 0
        
        for s in text_list:
            should, _ = self._should_anonymize(s)
            if not should: continue
            s = s.strip()
            if not s: continue

            if current_len + len(s) + len(DELIMITER) > MAX_CHUNK_SIZE and current_chunk:
                text_chunks.append(DELIMITER.join(current_chunk))
                current_chunk, current_len = [], 0
            
            current_chunk.append(s)
            current_len += len(s) + len(DELIMITER)
        
        if current_chunk:
            text_chunks.append(DELIMITER.join(current_chunk))

        if not text_chunks: return
        
        desc = f"Detecting Entities for {os.path.basename(self.file_path)}"
        for chunk in tqdm(text_chunks, desc=desc, leave=False):
            ner_records = self.orchestrator.detect_entities([chunk])
            for record in ner_records:
                if self.ner_file_handle is not None:
                    self.ner_file_handle.write(orjson.dumps(record).decode('utf-8') + "\n")
                    self.ner_file_handle.flush()


class TextFileProcessor(FileProcessor):
    def _get_output_extension(self) -> str:
        return ".txt"

    def _extract_texts(self) -> Iterable[str]:
        with open(self.file_path, "r", encoding="utf-8") as f:
            for line in f:
                yield line

    def _process_anonymization(self, output_path: str):
        """
        Custom anonymization processing for text files to handle newlines correctly.
        """
        with open(output_path, "w", encoding="utf-8") as outfile:
            # The base iterator yields lines with newlines attached.
            text_iterator = self._extract_texts()

            for text_batch in self._batch_iterator(text_iterator, self.DEFAULT_BATCH_SIZE):
                if not text_batch:
                    continue

                # The orchestrator expects clean text, so we strip trailing newlines.
                lines_to_process = [line.rstrip("\r\n") for line in text_batch]

                # Heuristic: Check the first line of the batch to decide if we should anonymize.
                # This assumes batches are reasonably homogenous.
                should_anonymize, forced_type = self._should_anonymize(lines_to_process[0])
                
                if should_anonymize:
                    anonymized_lines = self._process_batch_smart(lines_to_process, forced_entity_type=forced_type)
                    for line in anonymized_lines:
                        outfile.write(line + '\n')
                else:
                    # If the batch is skipped, write the original lines back.
                    for line in text_batch:
                        outfile.write(line)


class ImageFileProcessor(FileProcessor):
    def _get_output_extension(self) -> str:
        return ".txt"

    def _extract_texts(self) -> Iterable[str]:
        with open(self.file_path, "rb") as f:
            image_bytes = f.read()
        extracted_text = extract_text_from_image(image_bytes)
        if extracted_text:
            yield extracted_text


class DocxFileProcessor(FileProcessor):
    def _get_output_extension(self) -> str:
        return ".txt"

    def _extract_texts(self) -> Iterable[str]:
        doc = Document(self.file_path)
        desc = f"Extracting DOCX {os.path.basename(self.file_path)}"
        for para in tqdm(doc.paragraphs, desc=desc, unit="para", leave=False):
            para_content_parts = []
            for run in para.runs:
                if run._r.xpath(".//w:drawing"):
                    for r_id in run._r.xpath(".//@r:embed"):
                        try:
                            image_part = doc.part.related_parts[r_id]
                            ocr_text = extract_text_from_image(image_part.blob)
                            if ocr_text: para_content_parts.append(ocr_text)
                        except (KeyError, AttributeError):
                            continue
                else:
                    para_content_parts.append(run.text)
            full_para_text = "".join(para_content_parts)
            if full_para_text.strip():
                yield full_para_text + "\n"


class PdfFileProcessor(FileProcessor):
    def _get_output_extension(self) -> str:
        return ".txt"

    def _extract_texts(self) -> Iterable[str]:
        with fitz.open(self.file_path) as doc:
            desc = f"Extracting PDF {os.path.basename(self.file_path)}"
            for page_num in tqdm(range(doc.page_count), desc=desc, unit="page", leave=False):
                page = doc[page_num]
                logging.debug(f"Extracting text from page {page_num + 1}/{doc.page_count} of '{self.file_path}'.")
                content_items = []
                page_dict = page.get_text("dict")
                text_blocks = page_dict.get("blocks", []) if isinstance(page_dict, dict) else []
                for block in text_blocks:
                    if block["type"] == 0:
                        block_text = "".join(span["text"] for line in block.get("lines", []) for span in line.get("spans", []) if "text" in span)
                        if block_text.strip():
                            content_items.append({"bbox": block["bbox"], "content": block_text})

                for img in page.get_images(full=True):
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    if base_image:
                        bbox = page.get_image_bbox(img)
                        content_items.append({"bbox": bbox, "content": extract_text_from_image(base_image["image"])})

                content_items.sort(key=lambda item: (item["bbox"][1], item["bbox"][0]))
                
                for item in content_items:
                    if item['content']:
                        yield item['content'] + "\n"
                
                # Explicitly clean up PyMuPDF page objects to prevent memory leaks
                page.clean_contents()
                del page # Ensure the page object is released


class CsvFileProcessor(FileProcessor):
    def _get_output_extension(self) -> str:
        return ".csv"
    
    def _process_anonymization(self, output_path: str):
        chunk_size = self.csv_chunk_size
        temp_output_path = output_path + ".tmp"
        header_written = False
        
        use_deduplication = not self.preserve_row_context
        
        try:
            with open(self.file_path, "rb") as f:
                total_rows = sum(1 for _ in f) -1
        except (IOError, FileNotFoundError):
            total_rows = 0
        
        if total_rows <= 0:
            if not header_written:
                try:
                    df_header = pd.read_csv(self.file_path, nrows=0)
                    df_header.to_csv(output_path, mode='a', index=False, header=True)
                except Exception:
                    pass
            return

        try:
            with pd.read_csv(self.file_path, dtype=str, chunksize=chunk_size, on_bad_lines='warn', encoding='utf-8', low_memory=False) as reader:
                progress_bar = tqdm(total=total_rows, desc=f"Processing CSV {os.path.basename(self.file_path)}", unit="rows", leave=False)
                for chunk_idx, chunk in enumerate(reader):
                    logging.debug(f"Processing CSV chunk {chunk_idx + 1} with {len(chunk)} rows.")
                    anonymized_chunk = chunk.copy()
                    
                    if use_deduplication:
                        logging.debug("CSV processing with value deduplication (faster, context-unaware).")
                        texts_to_anonymize_map: Dict[Union[str, Tuple[str, ...]], Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))
                        for col in chunk.columns:
                            for val in chunk[col].dropna().unique():
                                val_str = str(val)
                                should_anon, forced_type = self._should_anonymize(val_str, col)
                                if should_anon:
                                    group_key = tuple(forced_type) if isinstance(forced_type, list) else (forced_type if forced_type is not None else "auto")
                                    texts_to_anonymize_map[group_key][col].append(val_str)

                        translation_map: Dict[str, str] = {}
                        for group_key, cols_data in texts_to_anonymize_map.items():
                            current_forced_type = group_key if group_key != "auto" else None
                            if isinstance(current_forced_type, tuple): current_forced_type = list(current_forced_type)
                            
                            unique_texts_for_group = list(set(val for sublist in cols_data.values() for val in sublist))
                            if unique_texts_for_group:
                                logging.debug(f"Anonymizing {len(unique_texts_for_group)} unique texts for group '{group_key}'.")
                                anonymized_texts = self._process_batch_smart(unique_texts_for_group, forced_entity_type=current_forced_type)
                                translation_map.update(dict(zip(unique_texts_for_group, anonymized_texts)))

                        if translation_map:
                            for col in chunk.columns:
                                logging.debug(f"Applying anonymization map to column '{col}'.")
                                anonymized_chunk[col] = anonymized_chunk[col].map(lambda x: translation_map.get(str(x), x))
                    else:
                        logging.debug("CSV processing with row context preservation (slower, context-aware).")
                        for col in chunk.columns:
                            series = chunk[col].dropna()
                            if series.empty: 
                                logging.debug(f"Column '{col}' is empty, skipping.")
                                continue
                            
                            if not any(self._should_anonymize(str(val), col)[0] for val in series):
                                logging.debug(f"Column '{col}' contains no PII to anonymize based on current rules.")
                                continue

                            values_to_process = series.tolist()
                            _, forced_type = self._should_anonymize(values_to_process[0], col)
                            
                            logging.debug(f"Processing {len(values_to_process)} values in column '{col}' for anonymization.")
                            anonymized_values = self._process_batch_smart(values_to_process, forced_entity_type=forced_type)
                            
                            anonymized_series = pd.Series(anonymized_values, index=series.index)
                            anonymized_chunk[col].update(anonymized_series)

                    anonymized_chunk.to_csv(temp_output_path, mode='a', index=False, header=not header_written, encoding='utf-8')
                    header_written = True
                    progress_bar.update(len(chunk))
                progress_bar.close()
            
            shutil.move(temp_output_path, output_path)
            logging.info(f"Successfully anonymized CSV file saved to: {output_path}")

        except Exception as e:
            logging.error(f"Error processing CSV file '{self.file_path}': {e}", exc_info=True)
            if os.path.exists(temp_output_path):
                os.remove(temp_output_path)
                logging.info(f"Removed corrupt temporary file: {temp_output_path}")
            raise # Re-raise the exception after cleanup

    def _extract_texts(self) -> Iterable[str]:
        chunk_size = self.csv_chunk_size
        with pd.read_csv(self.file_path, dtype=str, chunksize=chunk_size, on_bad_lines='warn', encoding='utf-8', low_memory=False) as reader:
            for chunk in reader:
                batch_values = []
                for col in chunk.columns:
                    for val in chunk[col].dropna():
                        val_str = str(val)
                        should_anon, _ = self._should_anonymize(val_str, col)
                        if should_anon:
                            batch_values.append(val_str)
                if batch_values:
                    yield "\n".join(batch_values) + "\n" # Yield batch of values separated by newline


class XlsxFileProcessor(FileProcessor):
    def _get_output_extension(self) -> str:
        return ".xlsx"

    def _extract_texts(self) -> Iterable[str]:
        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        path = f"{sheet.title}.{cell.column_letter}" # type: ignore
                        should_anon, _ = self._should_anonymize(cell.value, path)
                        if should_anon:
                            yield cell.value
    
    def _process_anonymization(self, output_path: str):
        """
        Process an XLSX file in a memory-efficient way.
        It makes two passes: one to collect texts and another to write the anonymized file.
        """
        # Pass 1: Collect all unique strings that need to be anonymized.
        try:
            read_only_wb = openpyxl.load_workbook(self.file_path, read_only=True)
        except Exception as e:
            logging.error(f"Failed to open XLSX file {self.file_path} in read-only mode: {e}")
            shutil.copy(self.file_path, output_path) # Copy original on failure to open
            return

        total_rows = sum(sheet.max_row for sheet in read_only_wb.worksheets)
        
        # The logic is now inverted: by default we use deduplication (faster).
        # Full context processing is enabled only when the flag is explicitly set.
        use_deduplication = not self.preserve_row_context
        logging.debug(f"XLSX processing pass 1 (collecting texts) with deduplication: {use_deduplication}.")
        
        all_texts_map: Dict[Union[str, Tuple[str, ...]], List[str]] = defaultdict(list)
        try:
            with tqdm(total=total_rows, desc=f"Pass 1/2: Reading {os.path.basename(self.file_path)}", unit="row", leave=False) as pbar:
                for sheet in read_only_wb.worksheets:
                    for row in sheet.iter_rows():
                        pbar.update(1)
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                path = f"{sheet.title}.{cell.column_letter}" # type: ignore
                                should_anon, forced_type = self._should_anonymize(cell.value, path)
                                if should_anon:
                                    group_key = tuple(forced_type) if isinstance(forced_type, list) else (forced_type or "auto")
                                    all_texts_map[group_key].append(cell.value)
        finally:
            read_only_wb.close() # Ensure read-only workbook is closed

        # Anonymize all collected texts at once.
        translation_map: Dict[str, deque] = defaultdict(deque)
        if all_texts_map:
            for group_key, texts in all_texts_map.items():
                forced_type = group_key if group_key != "auto" else None
                if isinstance(forced_type, tuple):
                    forced_type = list(forced_type)
                
                strings_to_process = sorted(list(set(texts))) if use_deduplication else texts
                if not strings_to_process: continue
                
                logging.debug(f"Anonymizing {len(strings_to_process)} texts for group '{group_key}' in XLSX.")
                anonymized_texts = self._process_batch_smart(strings_to_process, forced_entity_type=forced_type)
                for original, anonymized in zip(strings_to_process, anonymized_texts):
                    translation_map[original].append(anonymized)

        if not translation_map:
            logging.info("No PII found for anonymization in XLSX. Copying original file.")
            shutil.copy(self.file_path, output_path)
            return
            
        # Pass 2: Create a new workbook and write the anonymized content.
        try:
            read_only_wb = openpyxl.load_workbook(self.file_path, read_only=True)
        except Exception as e:
            logging.error(f"Failed to re-open XLSX file {self.file_path} for pass 2: {e}")
            # At this point, we can't create the anonymized file, so we copy the original.
            shutil.copy(self.file_path, output_path)
            return
            
        write_wb = openpyxl.Workbook()
        if "Sheet" in write_wb.sheetnames and len(write_wb.sheetnames) == 1:
            write_wb.remove(write_wb["Sheet"])

        logging.debug("XLSX processing pass 2 (writing anonymized file).")
        try:
            with tqdm(total=total_rows, desc=f"Pass 2/2: Writing {os.path.basename(self.file_path)}", unit="row", leave=False) as pbar:
                for read_sheet in read_only_wb.worksheets:
                    write_sheet = write_wb.create_sheet(title=read_sheet.title)
                    for row_idx, row in enumerate(read_sheet.iter_rows(), 1):
                        pbar.update(1)
                        for col_idx, cell in enumerate(row, 1):
                            new_cell = write_sheet.cell(row=row_idx, column=col_idx)
                            
                            original_value = cell.value
                            if isinstance(original_value, str) and original_value in translation_map:
                                try:
                                    if use_deduplication:
                                        new_cell.value = translation_map[original_value][0]
                                    else:
                                        new_cell.value = translation_map[original_value].popleft()
                                except IndexError:
                                    logging.error(f"Mismatch in anonymized XLSX cell counts for '{original_value}'. Using original value as fallback.")
                                    new_cell.value = original_value
                            else:
                                new_cell.value = original_value
        finally:
            read_only_wb.close()
        
        logging.info(f"Anonymized XLSX file saved to: {output_path}")
        write_wb.save(output_path)


class XmlFileProcessor(FileProcessor):
    XML_MEMORY_THRESHOLD_BYTES = 200 * 1024 * 1024 # 200 MB

    def _get_output_extension(self) -> str:
        return ".xml"

    def _get_xpath(self, elem) -> str:
        return "/".join(e.tag for e in elem.iterancestors()) + "/" + elem.tag

    def _extract_texts(self) -> Iterable[str]:
        for _, element in etree.iterparse(self.file_path, events=('end',), recover=True, strip_cdata=False):
            path = self._get_xpath(element)
            if element.text and element.text.strip():
                should_anon, _ = self._should_anonymize(element.text, path)
                if should_anon:
                    yield element.text
            if element.tail and element.tail.strip():
                should_anon, _ = self._should_anonymize(element.tail, path + "/tail()")
                if should_anon:
                    yield element.tail
            
            for key, value in element.attrib.items():
                attr_path = f"{path}[@{key}]"
                should_anon, _ = self._should_anonymize(value, attr_path)
                if should_anon:
                    yield value
            element.clear()

    def _process_anonymization(self, output_path: str):
        xml_memory_threshold_bytes = ProcessingLimits.XML_MEMORY_THRESHOLD_MB * 1024 * 1024
        file_size = os.path.getsize(self.file_path)
        if file_size > xml_memory_threshold_bytes and not self.force_large_xml:
            logging.error(
                f"XML file '{os.path.basename(self.file_path)}' ({file_size / 1024**2:.1f} MB) "
                f"exceeds the memory safety threshold of {ProcessingLimits.XML_MEMORY_THRESHOLD_MB} MB. "
                "Processing skipped to prevent potential Out-of-Memory errors and PII leakage. "
                "Use --force-large-xml to override and attempt processing, but be aware of memory implications and potential for data leakage if processing fails."
            )
            raise ValueError(
                f"XML file too large for safe processing without --force-large-xml. "
                f"Threshold: {ProcessingLimits.XML_MEMORY_THRESHOLD_MB} MB."
            )
        elif file_size > xml_memory_threshold_bytes and self.force_large_xml:
            logging.warning(
                f"Processing large XML file '{os.path.basename(self.file_path)}' ({file_size / 1024**2:.1f} MB) "
                "due to --force-large-xml flag. This may lead to high memory usage or Out-of-Memory errors, and PII leakage if processing fails."
            )

        parser = etree.XMLParser(recover=True, strip_cdata=False)
        try:
            tree = etree.parse(self.file_path, parser)
            if parser.error_log:
                for error in parser.error_log:
                    logging.warning(f"XML parsing error in {self.file_path} on line {error.line}: {error.message}")
        except etree.XMLSyntaxError as e:
            logging.error(f"Fatal XML syntax error in {self.file_path}: {e}", exc_info=True)
            with open(output_path, "w") as f:
                f.write(f"<!-- Could not parse XML file {os.path.basename(self.file_path)} due to syntax errors. -->")
            return

        use_deduplication = self.orchestrator.cache_manager.use_cache
        text_groups: Dict[Union[str, Tuple[str, ...]], List[str]] = defaultdict(list)
        logging.debug(f"XML processing with deduplication: {use_deduplication}.")

        # Wrap the iterator with tqdm for progress visualization
        nodes_iterator = tree.iter()
        total_nodes = len(list(tree.iter())) # This can be memory intensive for large files
        
        desc_collect = f"Pass 1/2: Collecting texts from {os.path.basename(self.file_path)}"
        for element in tqdm(tree.iter(), total=total_nodes, desc=desc_collect, unit="node", leave=False):
            path = self._get_xpath(element)
            if element.text and element.text.strip():
                should_anon, forced_type = self._should_anonymize(element.text, path)
                if should_anon:
                    group_key = tuple(forced_type) if isinstance(forced_type, list) else (forced_type if forced_type is not None else "auto")
                    text_groups[group_key].append(element.text)
                    logging.debug(f"Collected text for anonymization from path '{path}': '{element.text[:50]}...'")

            if element.tail and element.tail.strip():
                should_anon, forced_type = self._should_anonymize(element.tail, path + "/tail()")
                if should_anon:
                    group_key = tuple(forced_type) if isinstance(forced_type, list) else (forced_type if forced_type is not None else "auto")
                    text_groups[group_key].append(element.tail)
                    logging.debug(f"Collected tail text for anonymization from path '{path}/tail()': '{element.tail[:50]}...'")
            
            for key, value in element.attrib.items():
                attr_path = f"{path}[@{key}]"
                should_anon, forced_type = self._should_anonymize(value, attr_path)
                if should_anon:
                    group_key = tuple(forced_type) if isinstance(forced_type, list) else (forced_type if forced_type is not None else "auto")
                    text_groups[group_key].append(value)
                    logging.debug(f"Collected attribute value for anonymization from path '{attr_path}': '{value[:50]}...'")

        translation_map: Dict[str, deque] = defaultdict(deque)
        for group_key, texts in text_groups.items():
            forced_type = group_key if group_key != "auto" else None
            if isinstance(forced_type, tuple):
                forced_type = list(forced_type)
            
            strings_to_process = sorted(list(set(texts))) if use_deduplication else texts
            if not strings_to_process: continue

            logging.debug(f"Anonymizing {len(strings_to_process)} unique texts for group '{group_key}' in XML.")
            anonymized_texts = self._process_batch_smart(strings_to_process, forced_entity_type=forced_type)
            for original, anonymized in zip(strings_to_process, anonymized_texts):
                translation_map[original].append(anonymized)

        if not translation_map:
            logging.info("No PII found for anonymization in XML. Saving original file.")
            tree.write(output_path, encoding="utf-8", xml_declaration=True)
            return
            
        desc_anon = f"Pass 2/2: Anonymizing {os.path.basename(self.file_path)}"
        for element in tqdm(tree.iter(), total=total_nodes, desc=desc_anon, unit="node", leave=False):
            if element.text in translation_map:
                try:
                    if use_deduplication:
                        element.text = translation_map[element.text][0]
                    else:
                        element.text = translation_map[element.text].popleft()
                    logging.debug(f"Anonymized element text for '{element.tag}'.")
                except IndexError:
                    logging.error("Mismatch in anonymized XML text counts. Using original value as fallback.")

            if element.tail in translation_map:
                try:
                    if use_deduplication:
                        element.tail = translation_map[element.tail][0]
                    else:
                        element.tail = translation_map[element.tail].popleft()
                    logging.debug(f"Anonymized element tail for '{element.tag}'.")
                except IndexError:
                    logging.error("Mismatch in anonymized XML tail counts. Using original value as fallback.")
            
            for key, value in element.attrib.items():
                if value in translation_map:
                    try:
                        if use_deduplication:
                            element.set(key, translation_map[value][0])
                        else:
                            element.set(key, translation_map[value].popleft())
                        logging.debug(f"Anonymized attribute '{key}' for '{element.tag}'.")
                    except IndexError:
                        logging.error("Mismatch in anonymized XML attribute counts. Using original value as fallback.")

        logging.info(f"Anonymized XML file saved to: {output_path}")
        tree.write(output_path, encoding="utf-8", xml_declaration=True)


class JsonFileProcessor(FileProcessor):
    """
    Optimized JSON/JSONL processor that uses a hybrid approach.
    - For .jsonl files, it streams line by line.
    - For .json files, it uses a fast in-memory approach for small files and a
      memory-efficient streaming approach for large files that are root-level arrays.
    """
    def _get_output_extension(self) -> str:
        return ".json" if not self.file_path.endswith(".jsonl") else ".jsonl"

    def _is_json_array(self) -> bool:
        """Detects if the root of the JSON file is an array."""
        try:
            with open(self.file_path, 'rb') as f:
                for chunk in f:
                    stripped_chunk = chunk.lstrip()
                    if stripped_chunk:
                        return stripped_chunk.startswith(b'[')
            return False
        except (IOError, OSError):
            return False

    def _process_anonymization(self, output_path: str):
        """Dispatcher for JSON processing based on file type, structure, and size."""
        if self.file_path.endswith(".jsonl"):
            logging.info(f"Processing .jsonl file: {self.file_path}")
            self._process_anonymization_jsonl(output_path)
            return

        file_size = os.path.getsize(self.file_path)
        is_large_file = file_size >= self.json_stream_threshold_bytes
        logging.debug(f"JSON file size: {file_size / (1024 * 1024):.2f} MB. Large file threshold: {self.json_stream_threshold_bytes / (1024 * 1024):.2f} MB.")

        if is_large_file:
            if self._is_json_array():
                logging.info("Large JSON array detected (%.1f MB). Switching to memory-efficient array streaming mode.", file_size / 1024 / 1024)
                self._process_json_array_streaming(output_path, self.json_chunk_size)
            else:
                logging.error(f"Streaming for large single JSON objects ({file_size / 1024 / 1024:.1f} MB) is not supported due to memory safety. Only large arrays are streamable.")
                raise ValueError(f"Streaming for large single JSON objects ({file_size / 1024 / 1024:.1f} MB) is not supported due to memory safety. Only large arrays are streamable.")
        else:
            logging.info("Small JSON file detected. Processing in-memory.")
            self._process_anonymization_in_memory(output_path)

    def _process_json_array_streaming(self, output_path: str, chunk_size: int):
        """Processes a file containing a root-level JSON array in streamed chunks."""
        logging.info(f"Starting JSON array streaming for '{self.file_path}' with chunk size {chunk_size}.")
        temp_output_path = output_path + ".tmp"
        file_size = os.path.getsize(self.file_path)
        
        try:
            with open(self.file_path, 'rb') as in_f, \
                 open(temp_output_path, 'wb') as out_f, \
                 tqdm(total=file_size, unit='B', unit_scale=True, desc=f"Streaming {os.path.basename(self.file_path)}", leave=False) as pbar:
                
                out_f.write(b'[\n')
                is_first_chunk = True
                
                objects_iterator = ijson.items(in_f, 'item', use_float=True)
                
                last_pos = 0
                for chunk_idx, obj_batch in enumerate(self._batch_iterator(objects_iterator, chunk_size)):
                    current_pos = in_f.tell()
                    pbar.update(current_pos - last_pos)
                    last_pos = current_pos

                    logging.debug(f"Processing JSON array chunk {chunk_idx + 1} with {len(obj_batch)} objects.")
                    batch_text_groups = defaultdict(list)
                    
                    for obj in obj_batch:
                        obj_text_groups = self._collect_strings_from_object(obj)
                        for group_key, strings in obj_text_groups.items():
                            batch_text_groups[group_key].extend(strings)
                    
                    path_aware_map = self._build_path_aware_translation_map(batch_text_groups)
                    
                    for obj_idx, obj in enumerate(obj_batch):
                        # Use obj_idx to check if it's not the very first object in the very first chunk
                        if not (is_first_chunk and obj_idx == 0):
                            out_f.write(b',\n')
                        
                        reconstructed_obj = self._reconstruct_object(obj, path_aware_map)
                        out_f.write(orjson.dumps(reconstructed_obj, option=orjson.OPT_INDENT_2))
                    
                    is_first_chunk = False # After the first chunk, this is always false
            
                out_f.write(b'\n]') # Close the JSON array
            
            shutil.move(temp_output_path, output_path)
            logging.info(f"Finished JSON array streaming for '{self.file_path}'. Anonymized file saved to: {output_path}")

        except (ijson.JSONError, MemoryError) as e:
            logging.error(f"Error streaming JSON file {self.file_path}: {e}", exc_info=True)
            if os.path.exists(temp_output_path):
                os.remove(temp_output_path)
                logging.info(f"Removed corrupt temporary file: {temp_output_path}")
            raise # Re-raise the exception after cleanup


    def _process_anonymization_in_memory(self, output_path: str):
        """Fast in-memory processing for files smaller than the threshold."""
        logging.info(f"Starting in-memory JSON processing for '{self.file_path}'.")
        temp_output_path = output_path + ".tmp"
        try:
            with open(self.file_path, "rb") as f:
                try:
                    data = orjson.loads(f.read())
                except orjson.JSONDecodeError:
                    logging.error(f"Invalid JSON in {self.file_path}. Aborting.")
                    with open(temp_output_path, "wb") as out_f:
                        out_f.write(b"{}")
                    shutil.move(temp_output_path, output_path)
                    return

            text_groups = self._collect_strings_from_object(data)
            path_aware_map = self._build_path_aware_translation_map(text_groups)
            
            if not path_aware_map:
                logging.info("No PII found for anonymization in JSON. Copying original file.")
                shutil.copyfile(self.file_path, output_path)
                if os.path.exists(temp_output_path):
                    os.remove(temp_output_path) # Ensure temp file is removed if it was created
                return

            reconstructed_data = self._reconstruct_object(data, path_aware_map)
            with open(temp_output_path, "wb") as f:
                f.write(orjson.dumps(reconstructed_data, option=orjson.OPT_INDENT_2))
            
            shutil.move(temp_output_path, output_path)
            logging.info(f"In-memory JSON processing complete. Anonymized file saved to: {output_path}")

        except Exception as e:
            logging.error(f"Error processing JSON file '{self.file_path}': {e}", exc_info=True)
            if os.path.exists(temp_output_path):
                os.remove(temp_output_path)
                logging.info(f"Removed corrupt temporary file: {temp_output_path}")
            raise # Re-raise the exception after cleanup
    
    def _collect_strings_from_object(self, obj, path_prefix: str = "") -> Dict[str, List[str]]:
        text_groups = defaultdict(list)
        
        def _walk(sub_obj, current_path):
            if isinstance(sub_obj, dict):
                for k, v in sub_obj.items():
                    _walk(v, f"{current_path}.{k}")
            elif isinstance(sub_obj, list):
                for i, item in enumerate(sub_obj):
                    _walk(item, f"{current_path}[{i}]")
            elif isinstance(sub_obj, str):
                should_anon, forced_type = self._should_anonymize(sub_obj, current_path)
                if should_anon:
                    group_key = "auto"
                    if isinstance(forced_type, str):
                        group_key = forced_type
                    elif isinstance(forced_type, list):
                        group_key = tuple(sorted(forced_type))
                    
                    text_groups[group_key].append(sub_obj)

        _walk(obj, path_prefix.lstrip('.'))
        return text_groups
    
    def _reconstruct_object(self, obj, path_aware_map: Dict, path_prefix: str = ""):
        current_path = path_prefix.lstrip('.')
        if isinstance(obj, dict):
            return {k: self._reconstruct_object(v, path_aware_map, f"{current_path}.{k}") for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._reconstruct_object(item, path_aware_map, f"{current_path}[{i}]") for i, item in enumerate(obj)]
        elif isinstance(obj, str):
            should_anon, group_key_or_list = self._should_anonymize(obj, current_path)
            if not should_anon:
                return obj

            final_group_key: Union[str, tuple] = "auto"
            if isinstance(group_key_or_list, str):
                final_group_key = group_key_or_list
            elif isinstance(group_key_or_list, list):
                final_group_key = tuple(sorted(group_key_or_list))

            if final_group_key in path_aware_map and obj in path_aware_map[final_group_key]:
                try:
                    if self.orchestrator.cache_manager.use_cache:
                        # In cache mode, replacement is static (peek at the first item).
                        return path_aware_map[final_group_key][obj][0]
                    else:
                        # In full context mode, consume from the queue.
                        return path_aware_map[final_group_key][obj].popleft()
                except IndexError:
                    logging.error(f"Mismatch in anonymized string counts for string '{obj}' in group '{final_group_key}'. Using original value as fallback.")
                    return obj
        return obj

    def _build_path_aware_translation_map(self, text_groups: Dict) -> Dict[Union[str, tuple], Dict[str, 'deque']]:
        path_aware_map: Dict[Union[str, tuple], Dict[str, 'deque']] = defaultdict(lambda: defaultdict(deque))
        if not text_groups:
            logging.debug("No text groups found for translation map.")
            return path_aware_map

        use_deduplication = self.orchestrator.cache_manager.use_cache
        total_strings = sum(len(v) if not use_deduplication else len(set(v)) for v in text_groups.values())
        if not total_strings:
            logging.debug("No strings to process for translation map.")
            return path_aware_map
            
        progress_desc = "Anonymizing (cached)" if use_deduplication else "Anonymizing (full context)"
        logging.debug(f"Building path-aware translation map ({progress_desc}). Total strings to process: {total_strings}.")
        progress = tqdm(total=total_strings, desc=progress_desc, unit="str", leave=False)

        for group_key, string_list in text_groups.items():
            strings_to_process = sorted(list(set(string_list))) if use_deduplication else string_list
            if not strings_to_process: continue
            logging.debug(f"Processing group '{group_key}' with {len(strings_to_process)} strings.")

            forced_type: Optional[Union[str, List[str]]] = group_key if group_key != "auto" else None
            if isinstance(forced_type, tuple): 
                forced_type = list(forced_type)

            anonymized_strings = []
            for chunk in self._batch_iterator(strings_to_process, self.DEFAULT_BATCH_SIZE):
                anonymized_chunk = self._process_batch_smart(chunk, forced_entity_type=forced_type)
                anonymized_strings.extend(anonymized_chunk)
                progress.update(len(chunk))
            
            for original, anonymized in zip(strings_to_process, anonymized_strings):
                path_aware_map[group_key][original].append(anonymized)
        
        progress.close()
        logging.debug("Path-aware translation map built successfully.")
        return path_aware_map

    def _process_anonymization_jsonl(self, output_path: str):
        """Processes .jsonl files line by line for maximum memory efficiency."""
        temp_output_path = output_path + ".tmp"
        try:
            with open(self.file_path, "rb") as f:
                total_lines = sum(1 for _ in f)

            with open(temp_output_path, "wb") as out_f, open(self.file_path, "rb") as in_f:
                desc = f"Processing {os.path.basename(self.file_path)}"
                for line in tqdm(in_f, total=total_lines, desc=desc, unit="line", leave=False):
                    if not line.strip(): continue
                    try:
                        data = orjson.loads(line)
                        text_groups = self._collect_strings_from_object(data)
                        path_aware_map = self._build_path_aware_translation_map(text_groups)
                        processed_data = self._reconstruct_object(data, path_aware_map)
                        out_f.write(orjson.dumps(processed_data) + b'\n')
                    except orjson.JSONDecodeError:
                        logging.warning(f"Skipping invalid JSON line in {self.file_path}. Original line written.")
                        out_f.write(line) # Write original line if parsing fails
            
            shutil.move(temp_output_path, output_path)
            logging.info(f"Successfully anonymized JSONL file saved to: {output_path}")

        except Exception as e:
            logging.error(f"Error processing JSONL file '{self.file_path}': {e}", exc_info=True)
            if os.path.exists(temp_output_path):
                os.remove(temp_output_path)
                logging.info(f"Removed corrupt temporary file: {temp_output_path}")
            raise # Re-raise the exception after cleanup
    
    def _extract_texts(self) -> Iterable[str]:
        if self.file_path.endswith(".jsonl"):
             with open(self.file_path, "rb") as f:
                for line in f:
                    if not line.strip(): continue
                    try:
                        data = orjson.loads(line)
                        yield from self._yield_strings_from_obj(data)
                    except orjson.JSONDecodeError:
                        continue
        else:
            try:
                with open(self.file_path, 'rb') as f:
                    for obj in ijson.items(f, 'item', use_float=True, multiple_values=self._is_json_array()):
                        yield from self._yield_strings_from_obj(obj)
            except Exception as e:
                 logging.warning("Failed to stream JSON, falling back to in-memory loading. Error: %s", e)
                 with open(self.file_path, "rb") as f:
                    data = orjson.loads(f.read())
                    yield from self._yield_strings_from_obj(data)


    def _yield_strings_from_obj(self, obj, path=""):
        path = path.lstrip(".")
        if isinstance(obj, dict):
            for k, v in obj.items():
                yield from self._yield_strings_from_obj(v, f"{path}.{k}")
        elif isinstance(obj, list):
            for item in obj:
                yield from self._yield_strings_from_obj(item, path)
        elif isinstance(obj, str):
             should, _ = self._should_anonymize(obj, path)
             if should: yield obj



from .config import Global, ProcessingLimits, DefaultSizes
from .engine import AnonymizationOrchestrator

class ProcessorRegistry:
    """A registry for mapping file extensions to their corresponding processors."""
    _processors: Dict[str, type[FileProcessor]] = {}
    
    @classmethod
    def register(cls, extensions: List[str], processor_class: type[FileProcessor]):
        """Register a processor class for a list of file extensions."""
        for ext in extensions:
            cls._processors[ext.lower()] = processor_class

    @classmethod
    def get_processor(cls, file_path: str, orchestrator: AnonymizationOrchestrator, **kwargs) -> Optional[FileProcessor]:
        """
        Determines the correct processor for a file based on its file extension.
        """
        ext_to_use = os.path.splitext(file_path)[1].lower()

        processor_class = cls._processors.get(ext_to_use)
        if not processor_class:
            logging.debug(f"No suitable processor found for file type '{ext_to_use}' for file: '{file_path}'. Skipping.")
            return None

        logging.debug(f"Found processor '{processor_class.__name__}' for file type '{ext_to_use}'.")
        return processor_class(file_path, orchestrator, **kwargs)



# Register all the processors
ProcessorRegistry.register([".txt", ".log"], TextFileProcessor)
ProcessorRegistry.register([".pdf"], PdfFileProcessor)
ProcessorRegistry.register([".docx"], DocxFileProcessor)
ProcessorRegistry.register([".csv"], CsvFileProcessor)
ProcessorRegistry.register([".xlsx"], XlsxFileProcessor)
ProcessorRegistry.register([".xml"], XmlFileProcessor)
ProcessorRegistry.register([".json", ".jsonl"], JsonFileProcessor)
ProcessorRegistry.register([".jpeg", ".jpg", ".png", ".gif", ".bmp", ".tiff", ".tif", ".webp", ".jp2", ".pnm"], ImageFileProcessor)
