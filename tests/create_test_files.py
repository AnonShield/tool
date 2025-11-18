import fitz  # PyMuPDF
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from docx import Document
import os


def create_test_files(test_data_dir):
    os.makedirs(test_data_dir, exist_ok=True)
    email = "john.doe@gmail.com"
    name = "John Doe"
    location = "New York"
    phone = "555-1234"
    text = f"My name is {name} and my email is {email}. I live in {location} and my phone is {phone}."

    # --- Create Text-based files ---
    with open(os.path.join(test_data_dir, "test.txt"), "w") as f:
        f.write(text)

    with open(os.path.join(test_data_dir, "test.csv"), "w") as f:
        f.write(f"name,email,location,phone\n{name},{email},{location},{phone}\nJane Smith,jane.smith@example.com,London,555-5678")

    import json
    with open(os.path.join(test_data_dir, "test.json"), "w") as f:
        json.dump({"user": {"name": name, "email": email, "location": location, "phone": phone}}, f)

    with open(os.path.join(test_data_dir, "test.xml"), "w") as f:
        f.write(f"<user>\n  <name>{name}</name>\n  <email>{email}</email>\n  <location>{location}</location>\n  <phone>{phone}</phone>\n</user>")

    # --- Create PDF ---
    pdf_path = os.path.join(test_data_dir, "test.pdf")
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((50, 72), text)
    doc.save(pdf_path)
    doc.close()

    # --- Create DOCX ---
    docx_path = os.path.join(test_data_dir, "test.docx")
    doc = Document()
    doc.add_paragraph(text)
    doc.save(docx_path)

    # --- Create XLSX ---
    xlsx_path = os.path.join(test_data_dir, "test.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="name")
    ws.cell(row=1, column=2, value="email")
    ws.cell(row=1, column=3, value="location")
    ws.cell(row=1, column=4, value="phone")
    ws.cell(row=2, column=1, value=name)
    ws.cell(row=2, column=2, value=email)
    ws.cell(row=2, column=3, value=location)
    ws.cell(row=2, column=4, value=phone)
    wb.save(xlsx_path)

    # --- Create Image files ---
    img = Image.new('RGB', (1200, 200), color = (255, 255, 255))
    d = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("DejaVuSans.ttf", 40)
    except IOError:
        font = ImageFont.load_default()
    d.text((10,10), text, fill=(0,0,0), font=font)
    
    image_formats = {
        "png": "PNG",
        "jpeg": "JPEG",
        "gif": "GIF",
        "bmp": "BMP",
        "tiff": "TIFF",
        "webp": "WEBP",
        "jp2": "JPEG2000",
    }

    for ext, pillow_format in image_formats.items():
        try:
            img.save(os.path.join(test_data_dir, f"test.{ext}"), pillow_format)
        except (KeyError, ValueError):
            print(f"Warning: Pillow format {pillow_format} not supported for .{ext}")

if __name__ == "__main__":
    create_test_files("tests/test_data")
    create_pdf_with_images("tests/test_data/test_50_images.pdf", 50)