#! /usr/bin/env python
# /anon.py
import argparse
import io
import json
import os
import subprocess
import sys
import time
import warnings

import pymupdf as fitz
import numpy as np 
import openpyxl
import pandas as pd
import pytesseract
import spacy
import spacy.cli
from docx import Document
from huggingface_hub import snapshot_download
from lxml import etree # type: ignore
from PIL import Image

from config import DEFAULT_ALLOW_LIST, SECRET_KEY, TRANSFORMER_MODEL, TRF_MODEL_PATH
from engine import anonymize_text, get_presidio_engines, batch_process_text 

warnings.filterwarnings("ignore")

_ENGINES_CACHE = {}

# --- Dependency Checks ---
def tesseract_check():
    """Checks if Tesseract OCR is installed and accessible."""
    try:
        pytesseract.get_tesseract_version()
    except pytesseract.TesseractNotFoundError:
        print("[!] Tesseract is not installed or not in your PATH. It is required for processing images.", file=sys.stderr)
        sys.exit(1)

def ocr_check(file_path: str) -> bool:
    """Checks if a file requires OCR processing."""
    ext = os.path.splitext(file_path)[1].lower()
    return ext in (".pdf", ".docx", ".xlsx")


def ensure_ocr_policy(file_path: str, is_image_file: bool, need_ocr_hint: bool | None = None) -> bool:
    """Decide whether OCR should be performed for a given file

    - need_ocr_hint: if the caller already determined that OCR is
        needed (True/False), pass it to avoid re-opening the file
    """
    # If image file (PNG, JPG, etc), require Tesseract and abort early
    if is_image_file:
        tesseract_check()
        return True

    # Use caller's precomputed decision if available to avoid duplicate IO
    if need_ocr_hint is None:
        need_ocr = ocr_check(file_path)
    else:
        need_ocr = bool(need_ocr_hint)

    if not need_ocr:
        return False

    # if OCR is needed but Tesseract is missing, warn and skip OCR
    try:
        pytesseract.get_tesseract_version()
        return True
    except pytesseract.TesseractNotFoundError:
        print(f"[!] Warning: '{file_path}' contains images or scanned pages but Tesseract is not available. Image OCR will be skipped.", file=sys.stderr)
        return False

def models_check(lang: str):
    """Downloads and verifies necessary spaCy and Transformer models."""
    spacy_model_map = {"pt": "pt_core_news_lg", "en": "en_core_web_lg"}
    en_model = spacy_model_map["en"]

    # Try to download the requested spaCy model if not present
    requested = spacy_model_map.get(lang) or f"{lang}_core_news_lg"

    for model in (en_model, requested):
        if model and not spacy.util.is_package(model):
            print(f"[+] Spacy model '{model}' not found. Downloading...")
            try:
                subprocess.run(
                    [sys.executable, "-m", "spacy", "download", model],
                    check=True, capture_output=True, text=True,
                )
                print(f"[*] Successfully downloaded '{model}'.")
                if not spacy.util.is_package(model):
                    raise Exception(f"Model '{model}' downloaded but still not available.")
            except subprocess.CalledProcessError:
                print(f"[!] Failed to download spaCy model '{model}'.", file=sys.stderr)
                sys.exit(1)

    if not os.path.exists(TRF_MODEL_PATH):
        print(f"[!] Downloading Transformer model '{TRANSFORMER_MODEL}'...")
        snapshot_download(repo_id=TRANSFORMER_MODEL, cache_dir=TRF_MODEL_PATH, max_workers=10)

# --- Text Extraction ---
def extract_text_from_image(image_bytes: bytes) -> str:
    """Extracts text from image bytes using OCR."""
    try:
        return pytesseract.image_to_string(Image.open(io.BytesIO(image_bytes)))
    except Exception:
        return ""

# --- Lazy Loading Engine Manager ---
def get_or_create_engines_lazily(lang: str):
    """Lazily initializes and retrieves NLP engines from a cache."""
    if lang not in _ENGINES_CACHE:
        print(f"[+] Lazily initializing NLP engines for language '{lang}' (this may take a moment)...")
        _ENGINES_CACHE[lang] = get_presidio_engines(lang)
    return _ENGINES_CACHE[lang]

# --- File Processors ---

def process_plain_text_and_pdf(file_path, anonymizer_func):
    """Processes plain text and PDF files, including OCR for PDFs."""
    ext = os.path.splitext(file_path)[1].lower()
    content = ""
    if ext == ".txt":
        with open(file_path, "r", encoding="utf-8") as f: content = f.read()
    elif ext == ".pdf":
        parts: list[str] = []
        images_for_ocr: list[bytes] = []
        need_ocr = False

        with fitz.open(file_path) as doc:
            for page in doc:
                page_text = page.get_text()
                parts.append(page_text)

                images = page.get_images(full=True)
                if images:
                    # collect image bytes to OCR only if needed
                    for img in images:
                        base_image = doc.extract_image(img[0])
                        images_for_ocr.append(base_image["image"])
                    # if page has no extractable text, OCR would be needed
                    if not page_text.strip():
                        need_ocr = True

        if need_ocr and images_for_ocr:
            # decide via central policy whether to OCR for this document
            do_ocr = ensure_ocr_policy(file_path, is_image_file=False, need_ocr_hint=need_ocr)
            if do_ocr:
                for img_bytes in images_for_ocr:
                    parts.append(extract_text_from_image(img_bytes))

        content = "\n".join(parts)

    anonymized_content = anonymizer_func(content)
    output_path = get_output_path(file_path, ".txt")
    with open(output_path, "w", encoding="utf-8") as f: f.write(anonymized_content)
    return output_path

def process_docx_with_ocr(file_path, anonymizer_func):
    """Processes DOCX files, extracting text from paragraphs and images (OCR)."""
    doc = Document(file_path)
    data_parts = []
    images_to_process = []

    # Extract text and collect image blobs (don't OCR yet)
    for para in doc.paragraphs:
        data_parts.append(para.text)
        for run in para.runs:
            for inline in run._r.xpath(".//w:drawing"):
                blip_embeds = inline.xpath(".//a:blip/@r:embed")
                if blip_embeds:
                    for rel in doc.part.rels.values():
                        if "image" in rel.target_ref and rel.rId in blip_embeds[0]:
                            images_to_process.append(rel.target_part.blob)
                            data_parts.append("__IMAGE_PLACEHOLDER__")

    image_texts: list[str] = []
    if images_to_process:
        # We already detected embedded images; pass the precomputed flag to
        # avoid re-opening/parsing the DOCX inside ensure_ocr_policy.
        do_ocr = ensure_ocr_policy(file_path, is_image_file=False, need_ocr_hint=True)
        if do_ocr:
            image_texts = [extract_text_from_image(img_bytes) for img_bytes in images_to_process]

    # Combine text and image OCR results
    image_text_iter = iter(image_texts)
    full_content = "\n".join(
        [part if part != "__IMAGE_PLACEHOLDER__" else next(image_text_iter, "") for part in data_parts]
    )

    anonymized_content = anonymizer_func(full_content)
    output_path = get_output_path(file_path, ".txt")
    with open(output_path, "w", encoding="utf-8") as f: f.write(anonymized_content)
    return output_path

def process_csv_batched(file_path, lang, allow_list, entities_to_preserve):
    """Processes CSV files in batches for performance."""
    df = pd.read_csv(file_path, dtype=str)
    analyzer_engine, anonymizer_engine = get_or_create_engines_lazily(lang)

    all_values = [str(val) if val is not None else "" for val in df.values.flatten().tolist()]

    anonymized_values = batch_process_text(
        all_values, analyzer_engine, anonymizer_engine, lang, allow_list, entities_to_preserve
    )

    anonymized_array = np.array(anonymized_values).reshape(df.shape)
    anonymized_df = pd.DataFrame(anonymized_array, columns=df.columns, index=df.index)

    output_path = get_output_path(file_path, ".csv")
    anonymized_df.to_csv(output_path, index=False, encoding="utf-8")
    return output_path

def process_xlsx(file_path, anonymizer_func):
    """
    Processes XLSX files by anonymizing text in cells and replacing images
    with their anonymized OCR text.
    """
    wb = openpyxl.load_workbook(file_path)

    for sheet in wb.worksheets:
        image_replacements = []
        if hasattr(sheet, '_images') and sheet._images: # type: ignore
            images_to_process = list(sheet._images) # type: ignore

            # We already know the sheet has embedded images; pass that as a
            # precomputed hint so ensure_ocr_policy doesn't reopen the file.
            do_ocr = ensure_ocr_policy(file_path, is_image_file=False, need_ocr_hint=True)
            if do_ocr:
                for image in images_to_process:
                    # extract image bytes, ensure the openpyxl image keeps its data
                    img_bytes = image._data()
                    image._data = (lambda b: lambda: b)(img_bytes)

                    # Perform OCR per-image
                    ocr_text = extract_text_from_image(img_bytes)
                    if ocr_text.strip():
                        anonymized_ocr = anonymizer_func(ocr_text)
                        image_replacements.append((image.anchor, anonymized_ocr))

            sheet._images.clear() # type: ignore

        # Anonymize cell text
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell.value = anonymizer_func(cell.value)

        # Add the anonymized image text to cells
        for anchor, text in image_replacements:
            try:
                if hasattr(anchor, '_from') and hasattr(anchor._from, 'row') and hasattr(anchor._from, 'col'):
                    row = anchor._from.row + 1
                    col = anchor._from.col + 1
                    cell = sheet.cell(row=row, column=col)
                    
                    existing_value = cell.value
                    if existing_value:
                        cell.value = f"{text}\n{existing_value}" # type: ignore
                    else:
                        cell.value = text
                else:
                    # Fallback: append to end of sheet
                    sheet.append([f"Anonymized image text:", text])
            except Exception:
                # If any error occurs, append to end of sheet
                sheet.append([f"Anonymized image text:", text])

    output_path = get_output_path(file_path, ".xlsx")
    wb.save(output_path)
    return output_path

def process_xml(file_path, anonymizer_func):
    """Processes XML files, preserving their hierarchical structure."""
    parser = etree.XMLParser(recover=True, strip_cdata=False)
    tree = etree.parse(file_path, parser)
    for element in tree.iter():
        if element.text and element.text.strip(): element.text = anonymizer_func(element.text)
        if element.tail and element.tail.strip(): element.tail = anonymizer_func(element.tail)
    output_path = get_output_path(file_path, ".xml")
    tree.write(output_path, encoding="utf-8", xml_declaration=True)
    return output_path

def process_json(file_path, anonymizer_func):
    """Processes JSON files, preserving their nested structure."""
    with open(file_path, "r", encoding="utf-8") as f: data = json.load(f)

    def recursive_anonymize(obj):
        if isinstance(obj, dict): return {k: recursive_anonymize(v) for k, v in obj.items()}
        elif isinstance(obj, list): return [recursive_anonymize(item) for item in obj]
        elif isinstance(obj, str): return anonymizer_func(obj)
        return obj

    anonymized_data = recursive_anonymize(data)
    output_path = get_output_path(file_path, ".json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(anonymized_data, f, indent=4, ensure_ascii=False)
    return output_path

def process_image_file(file_path, anonymizer_func):
    """Processes a single image file using OCR."""
    # For standalone image files require Tesseract; ensure_ocr_policy will abort via tesseract_check()
    ensure_ocr_policy(file_path, is_image_file=True)

    with open(file_path, "rb") as f:
        image_bytes = f.read()

    extracted_text = extract_text_from_image(image_bytes)
    anonymized_content = anonymizer_func(extracted_text)

    output_path = get_output_path(file_path, ".txt")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(anonymized_content)
    return output_path

# --- Utility Functions ---
def get_output_path(original_path, new_ext):
    """Constructs the output file path in the 'output' directory."""
    os.makedirs("output", exist_ok=True)
    base_name = os.path.splitext(os.path.basename(original_path))[0]
    return os.path.join("output", f"anon_{base_name}{new_ext}")

def write_report(file_path, start_time):
    """Writes a simple performance report to the 'logs' directory."""
    os.makedirs("logs", exist_ok=True)
    base_name = os.path.basename(file_path)
    report_file = os.path.join("logs", f"report_{base_name}.txt")
    with open(report_file, "w", encoding="utf-8") as report:
        report.write(f"Processed file: {file_path}\n")
        report.write(f"Total elapsed time: {time.time() - start_time:.2f} seconds\n")
    print(f"Report saved at: {report_file}")

def main():
    """Main function to parse arguments and orchestrate the anonymization process."""
    parser = argparse.ArgumentParser(description="Anonymize sensitive information in various file formats.")
    parser.add_argument("file_path", help="Path to the file to be anonymized.")
    parser.add_argument("--preserve-entities", type=str, default="", help="Comma-separated list of entity types to preserve (e.g., 'LOCATION,ORGANIZATION').")
    parser.add_argument("--lang", type=str, default="en", choices=["ca", "zh", "hr", "da", "nl", "en", "fi", "fr", "de", "el", "it", "ja", "ko", "lt", "mk", "nb", "pl", "pt", "ro", "ru", "sl", "es", "sv", "uk"], help="Language of the document for model selection.")
    parser.add_argument("--allow-list", type=str, default="", help="Comma-separated list of terms to add to the allow list.")
    args = parser.parse_args()

    if not SECRET_KEY:
        print("[!] Error: ANON_SECRET_KEY environment variable not set.", file=sys.stderr)
        sys.exit(1)

    start_time = time.time()
    models_check(args.lang)

    allow_list = DEFAULT_ALLOW_LIST + [term.strip() for term in args.allow_list.split(',') if term]
    entities_to_preserve = [e.strip().upper() for e in args.preserve_entities.split(',') if e]

    def anonymizer_func(text_to_anonymize):
        analyzer_engine, anonymizer_engine = get_or_create_engines_lazily(args.lang)
        return anonymize_text(text_to_anonymize, analyzer_engine, anonymizer_engine, allow_list, entities_to_preserve, lang=args.lang)

    print(f"[+] Processing file: {args.file_path}...")
    ext = os.path.splitext(args.file_path)[1].lower()

    file_processors = {
        ".txt": process_plain_text_and_pdf,
        ".pdf": process_plain_text_and_pdf,
        ".docx": process_docx_with_ocr,
        ".csv": process_csv_batched,
        ".xlsx": process_xlsx,
        ".xml": process_xml,
        ".json": process_json,
        ".jpeg": process_image_file,
        ".png": process_image_file,
        ".gif": process_image_file,
        ".bmp": process_image_file,
        ".tiff": process_image_file,
        ".tif": process_image_file,
        ".webp": process_image_file,
        ".jp2": process_image_file,
        ".pnm": process_image_file,
    }

    processor = file_processors.get(ext)
    if not processor:
        print(f"[!] Unsupported file format: {ext}", file=sys.stderr)
        sys.exit(1)

    try:
        if ext == ".csv":
            output_file = processor(args.file_path, args.lang, allow_list, entities_to_preserve)
        else:
            output_file = processor(args.file_path, anonymizer_func)

        print(f"[*] Anonymized file saved at: {output_file}")
        write_report(args.file_path, start_time)
    except Exception as e:
        print(f"[!] An error occurred during processing: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()